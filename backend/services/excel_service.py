import os
import re
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as XLImage

from utils.column_detector import ColumnDetector
from utils.mechanical_requirements import search_mechanical_specified
from utils.template_fields import CHEMICAL_ELEMENTS, MECHANICAL_FIELDS
from config import resolve_mechanical_requirements_path

BASE_DIR = Path(__file__).resolve().parent.parent


def _cell_to_display_str(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (ValueError, TypeError):
        pass
    return str(value).strip()


CHEMICAL_SPECS = {
    '400/18': {'C':'3.60 - 3.80','Si':'2.00 - 2.25','Mn':'0.45 Max.','P':'0.050 Max.','S':'0.012 Max.','Cu':'0.200 Max.','Ni':'1.00 Max.','Mg':'0.020 - 0.060'},
    '400/18 LT': {'C':'3.60 - 3.80','Si':'2.00 - 2.25','Mn':'0.45 Max.','P':'0.050 Max.','S':'0.012 Max.','Cu':'0.200 Max.','Ni':'1.00 Max.','Mg':'0.020 - 0.060'},
    'EN-GJS-400-18U-LT': {'C':'3.60 - 3.80','Si':'2.00 - 2.25','Mn':'0.45 Max.','P':'0.050 Max.','S':'0.012 Max.','Cu':'0.200 Max.','Ni':'1.00 Max.','Mg':'0.020 - 0.060'},
    '500/7': {'C':'3.50 - 3.80','Si':'1.80 - 2.40','Mn':'0.50 Max.','P':'0.050 Max.','S':'0.012 Max.','Cu':'0.200 Max.','Ni':'1.00 Max.','Mg':'0.020 - 0.060'},
    '600/3': {'C':'3.50 - 3.80','Si':'2.20 - 2.80','Mn':'0.50 Max.','P':'0.050 Max.','S':'0.012 Max.','Cu':'0.200 Max.','Ni':'1.00 Max.','Mg':'0.020 - 0.060'},
}

def get_chemical_specified(grade):
    if not grade:
        return {}
    grade_clean = str(grade).strip()
    for key in CHEMICAL_SPECS:
        if key.lower() in grade_clean.lower() or grade_clean.lower() in key.lower():
            return CHEMICAL_SPECS[key]
    return {}

CHEMICAL_ORDER = ["C", "Si", "Mn", "P", "S", "Cu", "Ni", "Mg"]

MECH_CELL_MAP = {
    "tensile": ("D27", "D29"),
    "proof_stress": ("E27", "E29"),
    "elongation": ("F27", "F29"),
    "hardness_bhn": ("H27", "H29"),
    "impact_individual": ("J27", "J29"),
    "impact_mean": ("L27", "L29"),
}

# Template cell map (only .value is written — images/formatting preserved)
BASIC_CELLS = {
    "customer": "D6",
    "material_grade": "J6",
    "casting_name": "D8",
    "drawing_no": "J8",
    "heat_no": "D10",
    "casting_sl_no": "K10",
    "invoice_no_date": "D12",
    "doc_ref": "L2",
    "issue_no_dt": "L3",
    "rev_no_dt": "L4",
}

_MECH_LABEL_TO_KEY = {
    "tensile strength": "tensile",
    "0.2% proof stress": "proof_stress",
    "% elongation": "elongation",
    "hardness bhn": "hardness_bhn",
    "impact individual (j)": "impact_individual",
    "impact mean (j)": "impact_mean",
}


def _norm_heat(s: str) -> str:
    return re.sub(r"\s+", "", (s or "").upper())


def _norm_item(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())

def _replace_cover_logo(sheet, logo_path: str) -> None:
    if not os.path.isfile(logo_path):
        return
    try:
        if getattr(sheet, "_images", None):
            # Replace the top-left cover/logo picture with the local company logo.
            sheet._images.pop(0)
        logo = XLImage(logo_path)
        logo.width = 270
        logo.height = 237
        sheet.add_image(logo, "B2")
    except Exception as exc:
        print(f"Unable to replace cover logo: {exc}")

def _clear_cover_right_block(sheet) -> None:
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for row in range(2, 5):
        for col in range(4, 13):  # D:L
            sheet.cell(row=row, column=col).fill = white_fill
        for col in range(11, 13):  # K:L stays blank in the top-right corner
            sheet.cell(row=row, column=col).value = None



class ExcelService:
    def list_excel_columns(self, filepath: str) -> Dict[str, Any]:
        """Return header row index and column names from the metallurgy sheet."""
        det = ColumnDetector()
        _, hr = det.detect(filepath)
        df = pd.read_excel(filepath, sheet_name=0, header=hr, dtype=object)
        columns = [str(c).strip() for c in df.columns]
        return {"header_row": hr, "columns": columns}

    def suggested_mapping_from_detect(self, filepath: str) -> Dict[str, str]:
        """Auto-detected field_id -> Excel column name mapping."""
        det = ColumnDetector()
        cmap, hr = det.detect(filepath)
        df = pd.read_excel(filepath, sheet_name=0, header=hr, dtype=object)
        columns = [str(c).strip() for c in df.columns]
        mapping: Dict[str, str] = {}

        basic_keys = [
            "heat_no",
            "casting_name",
            "customer",
            "material_grade",
            "drawing_no",
            "casting_sl_no",
            "invoice_no_date",
            "doc_ref",
            "issue_no_dt",
            "rev_no_dt",
        ]
        for key in basic_keys:
            idx = cmap.get(key)
            if idx is not None and idx < len(columns):
                mapping[key] = columns[idx]

        for el, idx in (cmap.get("chemical") or {}).items():
            if idx is not None and idx < len(columns):
                mapping[f"chem_actual_{el}"] = columns[idx]
        for el, idx in (cmap.get("specified_chemical") or {}).items():
            if idx is not None and idx < len(columns):
                mapping[f"chem_spec_{el}"] = columns[idx]
        for mk, idx in (cmap.get("mechanical") or {}).items():
            if idx is not None and idx < len(columns):
                mapping[f"mech_actual_{mk}"] = columns[idx]
        for mk, idx in (cmap.get("specified_mechanical") or {}).items():
            if idx is not None and idx < len(columns):
                mapping[f"mech_spec_{mk}"] = columns[idx]
        return mapping

    def _cmap_from_column_mapping(self, df: pd.DataFrame, mapping: Dict[str, str]) -> Dict[str, Any]:
        """Build internal column index map from user field_id -> Excel column name."""
        columns = [str(c).strip() for c in df.columns]
        name_to_idx = {name: i for i, name in enumerate(columns)}
        cmap: Dict[str, Any] = {
            "chemical": {},
            "mechanical": {},
            "specified_chemical": {},
            "specified_mechanical": {},
        }
        basic_keys = [
            "heat_no",
            "casting_name",
            "customer",
            "material_grade",
            "drawing_no",
            "casting_sl_no",
            "invoice_no_date",
            "doc_ref",
            "issue_no_dt",
            "rev_no_dt",
        ]
        for key in basic_keys:
            col = (mapping.get(key) or "").strip()
            if col and col in name_to_idx:
                cmap[key] = name_to_idx[col]
        for el in CHEMICAL_ELEMENTS:
            col = (mapping.get(f"chem_actual_{el}") or "").strip()
            if col and col in name_to_idx:
                cmap["chemical"][el] = name_to_idx[col]
            col = (mapping.get(f"chem_spec_{el}") or "").strip()
            if col and col in name_to_idx:
                cmap["specified_chemical"][el] = name_to_idx[col]
        for m in MECHANICAL_FIELDS:
            mk = m["key"]
            col = (mapping.get(f"mech_actual_{mk}") or "").strip()
            if col and col in name_to_idx:
                cmap["mechanical"][mk] = name_to_idx[col]
            col = (mapping.get(f"mech_spec_{mk}") or "").strip()
            if col and col in name_to_idx:
                cmap["specified_mechanical"][mk] = name_to_idx[col]
        return cmap

    def _find_matching_row_with_cmap(
        self,
        df: pd.DataFrame,
        cmap: Dict[str, Any],
        heat_no: str,
        item_name: str,
    ) -> Tuple[Optional[pd.Series], str]:
        hidx = cmap.get("heat_no")
        if hidx is None:
            for i, c in enumerate(df.columns):
                if "heat" in str(c).lower():
                    hidx = i
                    break
        if hidx is None:
            return None, "No Heat No column mapped."

        iidx = cmap.get("casting_name")
        want_h = _norm_heat(heat_no)
        want_i = _norm_item(item_name)

        def _row_text(row: pd.Series) -> str:
            parts = []
            for col in df.columns:
                val = _cell_to_display_str(row[col])
                if val:
                    parts.append(val)
            return _norm_item(" ".join(parts))

        def _item_matches(item_cell: str) -> bool:
            if not want_i:
                return True
            if not item_cell:
                return False
            if want_i in item_cell or item_cell in want_i:
                return True
            tokens = [t for t in want_i.split() if len(t) > 2]
            return any(t in item_cell for t in tokens)

        for _, row in df.iterrows():
            cell_h = _norm_heat(self._row_value(row, df, hidx))
            if not cell_h:
                continue
            if want_h not in cell_h and cell_h not in want_h and cell_h != want_h:
                continue
            item_cell = _norm_item(self._row_value(row, df, iidx)) if iidx is not None else ""
            if not _item_matches(item_cell):
                continue
            return row, ""

        for _, row in df.iterrows():
            row_text = _row_text(row)
            if want_h and want_h not in row_text:
                continue
            if not _item_matches(row_text):
                continue
            return row, ""

        return None, f"No row matched Heat '{heat_no}' and Casting '{item_name}'."
    def search_mechanical_specified(
        self,
        casting_name: str,
        material_grade: str = "",
        mechanical_requirements_filename: Optional[str] = None,
    ) -> Dict[str, str]:
        return search_mechanical_specified(casting_name)

    def preview_with_mapping(
        self,
        filepath: str,
        heat_no: str,
        casting_name: str,
        column_mapping: Dict[str, str],
        spec_path: Optional[str] = None,
        mechanical_requirements_filename: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Load row data using user column mapping; always returns full field shell."""
        det = ColumnDetector()
        cmap, hr = det.detect(filepath)
        df = pd.read_excel(filepath, sheet_name=0, header=hr, dtype=object)
        if column_mapping:
            cmap = self._cmap_from_column_mapping(df, column_mapping)

        row, err = self._find_matching_row_with_cmap(df, cmap, heat_no, casting_name)
        row_data: Dict[str, Any]
        if row is not None:
            row_data = self._row_to_payload(row, df, cmap, heat_no)
        else:
            row_data = {
                "found": False,
                "message": err or "No matching row.",
                "customer": "",
                "material_grade": "",
                "drawing_no": "",
                "invoice_no_date": "",
                "casting_name": casting_name,
                "casting_sl_no": "",
                "heat_no": heat_no,
                "doc_ref": "",
                "issue_no_dt": "",
                "rev_no_dt": "",
                "chemical_actual": {el: "" for el in CHEMICAL_ORDER},
                "mechanical_actual": {k: "" for k in MECH_CELL_MAP},
            }

        spec_block = None
        if spec_path and os.path.isfile(spec_path):
            try:
                spec_block = self.search_specification_preview(spec_path, heat_no, casting_name)
            except Exception:
                spec_block = None

        resolved_casting = row_data.get("casting_name") or casting_name
        resolved_grade = row_data.get("material_grade") or ""
        mech_specified = self.search_mechanical_specified(
            resolved_casting,
            resolved_grade,
            mechanical_requirements_filename,
        )

        api = self.format_search_api_response(row_data, spec_block, mechanical_requirements=mech_specified)
        api["row_found"] = row is not None
        api["message"] = row_data.get("message", "") if row is None else "Data loaded."
        return api

    def _row_value(self, row: pd.Series, df: pd.DataFrame, col_idx: Optional[int]) -> str:
        if col_idx is None or col_idx >= len(df.columns):
            return ""
        return _cell_to_display_str(row[df.columns[col_idx]])

    def _find_matching_row(
        self,
        filepath: str,
        heat_no: str,
        item_name: str,
        column_mapping: Optional[Dict[str, str]] = None,
    ) -> Tuple[Optional[pd.Series], Optional[pd.DataFrame], Dict[str, Any], str]:
        det = ColumnDetector()
        cmap, hr = det.detect(filepath)
        df = pd.read_excel(filepath, sheet_name=0, header=hr, dtype=object)
        if column_mapping:
            cmap = self._cmap_from_column_mapping(df, column_mapping)

        row, err = self._find_matching_row_with_cmap(df, cmap, heat_no, item_name)
        if row is None:
            return None, df, cmap, err
        return row, df, cmap, ""

    def search_heat_no(self, filepath: str, heat_no: str, item_name: str = "") -> Dict[str, Any]:
        """Find row by heat (and optional item) for preview / legacy UI."""
        row, df, cmap, err = self._find_matching_row(filepath, heat_no, item_name or "")
        if row is None:
            return {"found": False, "message": err or "Not found."}
        return self._row_to_payload(row, df, cmap, heat_no)

    def search_specification_preview(
        self,
        spec_path: str,
        heat_no: str,
        item_name: str = "",
    ) -> Optional[Dict[str, Any]]:
        """
        Optional limits from a specification workbook. Returns None if file is missing,
        no row matches, or anything fails — callers treat None as 'no spec data'.
        """
        try:
            if not spec_path or not os.path.isfile(spec_path):
                return None
            row_s, df_s, cmap_s, err_s = self._find_matching_row(spec_path, heat_no, item_name or "")
            if row_s is None:
                return None
            schem = cmap_s.get("specified_chemical") or {}
            chem_plain = cmap_s.get("chemical") or {}
            chem_spec = {}
            for el in CHEMICAL_ORDER:
                v = self._row_value(row_s, df_s, schem.get(el))
                if not v:
                    v = self._row_value(row_s, df_s, chem_plain.get(el))
                chem_spec[el] = v
            smech = cmap_s.get("specified_mechanical") or {}
            mech_plain = cmap_s.get("mechanical") or {}
            mech_spec = {}
            for k in ["tensile", "proof_stress", "elongation", "hardness_bhn", "impact_individual", "impact_mean"]:
                v = self._row_value(row_s, df_s, smech.get(k))
                if not v:
                    v = self._row_value(row_s, df_s, mech_plain.get(k))
                mech_spec[k] = v
            return {"chemical": chem_spec, "mechanical": mech_spec}
        except Exception:
            return None

    def _row_to_payload(self, row: pd.Series, df: pd.DataFrame, cmap: Dict[str, Any], heat_no: str) -> Dict[str, Any]:
        ch = cmap.get("chemical") or {}
        mh = cmap.get("mechanical") or {}
        sch = cmap.get("specified_chemical") or {}
        smh = cmap.get("specified_mechanical") or {}
        chemical_actual = {el: self._row_value(row, df, ch.get(el)) for el in CHEMICAL_ORDER}
        chemical_specified = {el: self._row_value(row, df, sch.get(el)) for el in CHEMICAL_ORDER}
        mechanical_keys = ["tensile", "proof_stress", "elongation", "hardness_bhn", "impact_individual", "impact_mean"]
        mechanical_actual = {k: self._row_value(row, df, mh.get(k)) for k in mechanical_keys}
        mechanical_specified = {k: self._row_value(row, df, smh.get(k)) for k in mechanical_keys}
        casting = self._row_value(row, df, cmap.get("casting_name"))
        return {
            "found": True,
            "message": "Match found.",
            "customer": self._row_value(row, df, cmap.get("customer")),
            "material_grade": self._row_value(row, df, cmap.get("material_grade")),
            "drawing_no": self._row_value(row, df, cmap.get("drawing_no")),
            "invoice_no_date": self._row_value(row, df, cmap.get("invoice_no_date")),
            "item_name": casting,
            "casting_name": casting,
            "casting_sl_no": self._row_value(row, df, cmap.get("casting_sl_no")),
            "heat_no": self._row_value(row, df, cmap.get("heat_no")) or heat_no,
            "doc_ref": self._row_value(row, df, cmap.get("doc_ref")),
            "issue_no_dt": self._row_value(row, df, cmap.get("issue_no_dt")),
            "rev_no_dt": self._row_value(row, df, cmap.get("rev_no_dt")),
            "chemical_actual": chemical_actual,
            "chemical_specified": chemical_specified,
            "mechanical_actual": mechanical_actual,
            "mechanical_specified": mechanical_specified,
        }

    def format_search_api_response(
        self,
        row_data: Dict[str, Any],
        spec_block: Optional[Dict[str, Any]],
        mechanical_requirements: Optional[Dict[str, str]] = None,
    ) -> Dict[str, Any]:
        # Chemical specified: optional spec sheet / metallurgy columns (no separate chemical req file yet)
        chem_spec = (spec_block or {}).get("chemical") if spec_block else row_data.get("chemical_specified")
        if not chem_spec:
            chem_spec = {}

        # Mechanical specified: Mechanical_properties_Requriment.xlsx takes priority
        if mechanical_requirements is not None:
            mech_spec = mechanical_requirements
        elif spec_block and spec_block.get("mechanical"):
            mech_spec = spec_block["mechanical"]
        else:
            mech_spec = row_data.get("mechanical_specified") or {}
        return {
            "success": True,
            "basic_info": {
                "customer": row_data.get("customer", ""),
                "material_grade": row_data.get("material_grade", ""),
                "drawing_no": row_data.get("drawing_no", ""),
                "casting_sl_no": row_data.get("casting_sl_no", ""),
                "invoice_no_date": row_data.get("invoice_no_date", ""),
                "heat_no": row_data.get("heat_no", ""),
                "casting_name": row_data.get("casting_name") or row_data.get("item_name", ""),
                "doc_ref": row_data.get("doc_ref", ""),
                "issue_no_dt": row_data.get("issue_no_dt", ""),
                "rev_no_dt": row_data.get("rev_no_dt", ""),
            },
            "chemical_actual": row_data.get("chemical_actual") or {},
            "chemical_specified": chem_spec,
            "mechanical_actual": row_data.get("mechanical_actual") or {},
            "mechanical_specified": mech_spec,
        }

    def _dicts_to_table_rows(
        self,
        chem_act: Dict[str, str],
        chem_spec: Dict[str, str],
        mech_act: Dict[str, str],
        mech_spec: Dict[str, str],
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        chemical = [
            {
                "element": el,
                "specified": (chem_spec or {}).get(el, ""),
                "actual": (chem_act or {}).get(el, ""),
            }
            for el in CHEMICAL_ORDER
        ]
        mechanical = []
        for key in ["tensile", "proof_stress", "elongation", "hardness_bhn", "impact_individual", "impact_mean"]:
            label = key.replace("_", " ").title()
            if key == "proof_stress":
                label = "0.2% Proof Stress"
            elif key == "elongation":
                label = "% Elongation"
            elif key == "hardness_bhn":
                label = "Hardness BHN"
            elif key == "impact_individual":
                label = "Impact Individual (J)"
            elif key == "impact_mean":
                label = "Impact Mean (J)"
            mechanical.append(
                {
                    "key": key,
                    "name": label,
                    "specified": (mech_spec or {}).get(key, ""),
                    "actual": (mech_act or {}).get(key, ""),
                }
            )
        return chemical, mechanical

    def build_fill_payload(
        self,
        actual_path: str,
        spec_path: str,
        heat_no: str,
        item_name: str,
        user_heat: str,
        user_item: str,
    ) -> Dict[str, Any]:
        """Merge actual + specification rows for template writing."""
        row_a, df_a, cmap_a, err_a = self._find_matching_row(actual_path, heat_no, item_name)
        if row_a is None:
            return {"ok": False, "message": f"Actual sheet: {err_a}"}

        row_s, df_s, cmap_s, err_s = self._find_matching_row(spec_path, heat_no, item_name)
        if row_s is None:
            return {"ok": False, "message": f"Specification sheet: {err_s}"}

        customer = self._row_value(row_a, df_a, cmap_a.get("customer"))
        casting_sl = self._row_value(row_a, df_a, cmap_a.get("casting_sl_no"))
        heat_cell = self._row_value(row_a, df_a, cmap_a.get("heat_no")) or user_heat
        item_cell = self._row_value(row_a, df_a, cmap_a.get("casting_name")) or user_item

        chem_act = {el: self._row_value(row_a, df_a, cmap_a["chemical"].get(el)) for el in CHEMICAL_ORDER}
        mech_act = {
            k: self._row_value(row_a, df_a, cmap_a["mechanical"].get(k))
            for k in ["tensile", "proof_stress", "elongation", "hardness_bhn", "impact_individual", "impact_mean"]
        }

        # Specified: prefer specified_* maps; fall back to same mechanical/chemical columns
        chem_spec = {}
        for el in CHEMICAL_ORDER:
            v = self._row_value(row_s, df_s, cmap_s["specified_chemical"].get(el))
            if not v:
                v = self._row_value(row_s, df_s, cmap_s["chemical"].get(el))
            chem_spec[el] = v

        mech_spec = {}
        for k in ["tensile", "proof_stress", "elongation", "hardness_bhn", "impact_individual", "impact_mean"]:
            v = self._row_value(row_s, df_s, cmap_s["specified_mechanical"].get(k))
            if not v:
                v = self._row_value(row_s, df_s, cmap_s["mechanical"].get(k))
            mech_spec[k] = v

        return {
            "ok": True,
            "message": "Merged.",
            "d6_customer": customer,
            "d8_item": user_item.strip() or item_cell,
            "d10_heat": user_heat.strip() or heat_cell,
            "k10_casting_sl": casting_sl,
            "chemical_actual": chem_act,
            "chemical_spec": chem_spec,
            "mechanical_actual": mech_act,
            "mechanical_spec": mech_spec,
        }

    def _set_cell_value_only(self, ws, address: Optional[str], value: Optional[str]) -> None:
        """Assign .value only so number formats / borders / merges on the template stay intact."""
        if not address:
            return
        if value is None or str(value).strip() == "":
            return
        value_text = str(value).strip()
        merged_range = None
        for merged in ws.merged_cells.ranges:
            if address in merged:
                merged_range = merged
                break

        if merged_range:
            min_col = merged_range.min_col
            min_row = merged_range.min_row
            anchor = ws.cell(row=min_row, column=min_col)
            ws.unmerge_cells(str(merged_range))
            anchor.value = value_text
            ws.merge_cells(str(merged_range))
            return

        ws[address].value = value_text

    def _blank_report_workbook(self) -> Any:
        wb = Workbook()
        ws = wb.active
        ws.title = "Material Test Report"
        return wb

    def _mechanical_key_from_row(self, row: Dict[str, Any]) -> Optional[str]:
        k = _cell_to_display_str(row.get("key"))
        if k in MECH_CELL_MAP:
            return k
        name = _cell_to_display_str(row.get("name")).lower()
        return _MECH_LABEL_TO_KEY.get(name)

    def generate_report(
        self,
        template_filename: str,
        metallurgy_actual_filename: str,
        heat_no: str,
        casting_name: str,
        customer: str = "",
        material_grade: str = "",
        drawing_no: str = "",
        casting_sl_no: str = "",
        invoice_no_date: str = "",
        doc_ref: str = "",
        issue_no_dt: str = "",
        rev_no_dt: str = "",
        chemical_rows: Optional[List[Dict[str, Any]]] = None,
        mechanical_rows: Optional[List[Dict[str, Any]]] = None,
        chemical_actual: Optional[Dict[str, str]] = None,
        chemical_specified: Optional[Dict[str, str]] = None,
        mechanical_actual: Optional[Dict[str, str]] = None,
        mechanical_specified: Optional[Dict[str, str]] = None,
    ) -> Dict[str, str]:
        """Write the test report from preview data. load_workbook preserves template images."""
        actual_path = os.path.join("uploads", os.path.basename(metallurgy_actual_filename))
        if not os.path.isfile(actual_path):
            raise FileNotFoundError("Metallurgy actual file not found.")

        template_path = os.path.join("uploads", os.path.basename(template_filename))
        if not os.path.isfile(template_path):
            raise FileNotFoundError("Test Report template not found.")

        try:
            row_fallback = self.search_heat_no(actual_path, heat_no, casting_name)
        except Exception:
            row_fallback = {}

        if chemical_rows is None and chemical_actual is not None:
            fallback_chemical_actual = (row_fallback.get("chemical_actual") or {}) if isinstance(row_fallback, dict) else {}
            fallback_chemical_specified = (row_fallback.get("chemical_specified") or {}) if isinstance(row_fallback, dict) else {}
            fallback_mechanical_actual = (row_fallback.get("mechanical_actual") or {}) if isinstance(row_fallback, dict) else {}
            fallback_mechanical_specified = (row_fallback.get("mechanical_specified") or {}) if isinstance(row_fallback, dict) else {}

            merged_chemical_actual = {**fallback_chemical_actual, **(chemical_actual or {})}
            merged_chemical_specified = {**fallback_chemical_specified, **(chemical_specified or {})}
            merged_mechanical_actual = {**fallback_mechanical_actual, **(mechanical_actual or {})}
            merged_mechanical_specified = {**fallback_mechanical_specified, **(mechanical_specified or {})}

            for key, actual_value in list(merged_chemical_actual.items()):
                if not _cell_to_display_str(merged_chemical_specified.get(key)):
                    merged_chemical_specified[key] = actual_value
            for key, actual_value in list(merged_mechanical_actual.items()):
                if not _cell_to_display_str(merged_mechanical_specified.get(key)):
                    merged_mechanical_specified[key] = actual_value

            chemical_rows, mechanical_rows = self._dicts_to_table_rows(
                merged_chemical_actual,
                merged_chemical_specified,
                merged_mechanical_actual,
                merged_mechanical_specified,
            )
        chemical_rows = chemical_rows or []
        mechanical_rows = mechanical_rows or []

        wb = load_workbook(template_path, keep_links=True)
        ws = wb.active
        _replace_cover_logo(ws, os.path.join(str(BASE_DIR), "company_logo.png"))
        _clear_cover_right_block(ws)

        basic_fallback = row_fallback if isinstance(row_fallback, dict) else {}

        self._set_cell_value_only(ws, BASIC_CELLS["customer"], customer or basic_fallback.get("customer", ""))
        self._set_cell_value_only(ws, BASIC_CELLS["material_grade"], material_grade or basic_fallback.get("material_grade", ""))
        self._set_cell_value_only(ws, BASIC_CELLS["casting_name"], casting_name or basic_fallback.get("casting_name") or basic_fallback.get("item_name", ""))
        self._set_cell_value_only(ws, BASIC_CELLS["drawing_no"], drawing_no or basic_fallback.get("drawing_no", ""))
        self._set_cell_value_only(ws, BASIC_CELLS["heat_no"], heat_no or basic_fallback.get("heat_no", ""))
        self._set_cell_value_only(ws, BASIC_CELLS["casting_sl_no"], casting_sl_no or basic_fallback.get("casting_sl_no", ""))
        self._set_cell_value_only(ws, BASIC_CELLS["invoice_no_date"], invoice_no_date or basic_fallback.get("invoice_no_date", ""))

        chem_by_el: Dict[str, Dict[str, Any]] = {}
        for r in chemical_rows:
            el = _cell_to_display_str(r.get("element"))
            if el:
                chem_by_el[el] = r

        mech_vals: Dict[str, Tuple[str, str]] = {}
        for r in mechanical_rows:
            mk = self._mechanical_key_from_row(r)
            if not mk:
                continue
            spec = _cell_to_display_str(r.get("specified"))
            act = _cell_to_display_str(r.get("actual"))
            mech_vals[mk] = (spec, act)

        for j, el in enumerate(CHEMICAL_ORDER):
            col = 3 + j
            addr17 = ws.cell(row=17, column=col).coordinate
            addr19 = ws.cell(row=19, column=col).coordinate
            row = chem_by_el.get(el)
            spec_txt = _cell_to_display_str(row.get("specified")) if row else ""
            act_txt = _cell_to_display_str(row.get("actual")) if row else ""
            if spec_txt:
                # The template uses one merged cell per element in row 17-18, so
                # write the full spec string once instead of splitting it across rows.
                self._set_cell_value_only(ws, addr17, spec_txt)
            self._set_cell_value_only(ws, addr19, act_txt)
        for mk, (spec_cell, act_cell) in MECH_CELL_MAP.items():
            pair = mech_vals.get(mk)
            if not pair:
                continue
            self._set_cell_value_only(ws, spec_cell, pair[0])
            if mk == "impact_individual":
                impact_values = [v.strip() for v in str(pair[1]).split('\n') if v.strip()]
                self._set_cell_value_only(ws, "J29", impact_values[0] if len(impact_values) > 0 else "")
                self._set_cell_value_only(ws, "J30", impact_values[1] if len(impact_values) > 1 else "")
                self._set_cell_value_only(ws, "J31", impact_values[2] if len(impact_values) > 2 else "")
            else:
                self._set_cell_value_only(ws, act_cell, pair[1])

        out_name = f"MTR_{uuid.uuid4().hex[:10]}.xlsx"
        out_path = os.path.join("outputs", out_name)
        os.makedirs("outputs", exist_ok=True)
        wb.save(out_path)
        return {"output_filename": out_name, "output_path": out_path}

