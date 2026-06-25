import os
import re
import uuid
from pathlib import Path
from typing import Tuple

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as XLImage

from dependencies.auth import get_current_user
from models.schemas import DetectColumnsRequest, GenerateReportRequest, HeatSearchRequest, PreviewMappedRequest
from utils.template_fields import template_field_catalog
from services.excel_service import ExcelService
from utils.mechanical_requirements import search_mechanical_specified, search_chemical_specified_in_requirements
import pandas as pd

router = APIRouter(dependencies=[Depends(get_current_user)])
excel_service = ExcelService()

# Use absolute paths so they work regardless of working directory
BASE_DIR = Path(__file__).resolve().parent.parent
UPLOADS_DIR = str(BASE_DIR / "uploads")
OUTPUTS_DIR = str(BASE_DIR / "outputs")
os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(OUTPUTS_DIR, exist_ok=True)


def _latest_upload(prefix: str) -> str:
    if not os.path.isdir(UPLOADS_DIR):
        return ""
    matches = [
        name
        for name in os.listdir(UPLOADS_DIR)
        if name.startswith(prefix) and name.lower().endswith((".xlsx", ".xls"))
    ]
    if not matches:
        return ""
    matches.sort(key=lambda name: os.path.getmtime(os.path.join(UPLOADS_DIR, name)), reverse=True)
    return matches[0]


def _required_upload(filename: str, prefix: str, label: str) -> str:
    selected = os.path.basename((filename or "").strip()) or _latest_upload(prefix)
    if not selected or not os.path.isfile(os.path.join(UPLOADS_DIR, selected)):
        raise HTTPException(status_code=400, detail=f"{label} file not found.")
    return selected


def _store_upload(prefix: str, file: UploadFile) -> Tuple[str, str]:
    safe = os.path.basename(file.filename or f"{prefix}.xlsx")
    if not safe.lower().endswith((".xlsx", ".xls")):
        safe = f"{safe}.xlsx"
    stored = f"{prefix}_{uuid.uuid4().hex[:10]}_{safe}"
    dest = os.path.join(UPLOADS_DIR, stored)
    return stored, dest


@router.post("/upload_metallurgy")
@router.post("/upload/metallurgy")
async def upload_metallurgy(file: UploadFile = File(...)):
    stored, dest = _store_upload("metallurgy", file)
    content = await file.read()
    with open(dest, "wb") as f:
        f.write(content)
    return {
        "status": "success",
        "message": "Metallurgy (actual) uploaded.",
        "metallurgy_actual_filename": stored,
        "metallurgy_filename": stored,
    }


@router.post("/upload_template")
@router.post("/upload/template")
async def upload_template(file: UploadFile = File(...)):
    stored, dest = _store_upload("template", file)
    content = await file.read()
    with open(dest, "wb") as f:
        f.write(content)
    return {
        "status": "success",
        "message": "Template uploaded.",
        "template_filename": stored,
    }


def _casting_name(body: HeatSearchRequest) -> str:
    return (body.casting_name or body.item_name or "").strip()


def _mech_req_filename(body) -> str:
    return (getattr(body, "mechanical_requirements_filename", None) or "").strip()


def _replace_cover_logo(sheet, logo_path: str) -> None:
    if not os.path.isfile(logo_path):
        return
    try:
        if getattr(sheet, "_images", None):
            # Replace the top-left cover/logo picture with the local company logo.
            sheet._images.pop(0)
        logo = XLImage(logo_path)
        logo.width = 270
        logo.height = 237
        sheet.add_image(logo, "B2")
    except Exception as exc:
        print(f"Unable to replace cover logo: {exc}")

def _clear_cover_right_block(sheet) -> None:
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for row in range(2, 5):
        for col in range(4, 13):  # D:L
            sheet.cell(row=row, column=col).fill = white_fill
        for col in range(11, 13):  # K:L stays blank in the top-right corner
            sheet.cell(row=row, column=col).value = None



@router.post("/upload_mechanical_requirements")
@router.post("/upload_mechanical_specified")
@router.post("/upload/mechanical")
async def upload_mechanical_requirements(file: UploadFile = File(...)):
    stored, dest = _store_upload("mechanical", file)
    content = await file.read()
    with open(dest, "wb") as f:
        f.write(content)
    return {
        "status": "success",
        "message": "Mechanical properties requirements uploaded.",
        "mechanical_requirements_filename": stored,
    }

@router.get("/template_fields")
async def get_template_fields():
    return {"fields": template_field_catalog()}


@router.post("/detect_columns")
async def detect_columns(body: DetectColumnsRequest):
    fn = _required_upload(body.metallurgy_filename, "metallurgy", "Metallurgy")
    path = os.path.join(UPLOADS_DIR, fn)
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="Metallurgy file not found.")
    try:
        meta = excel_service.list_excel_columns(path)
        suggested = excel_service.suggested_mapping_from_detect(path)
        return {
            "status": "success",
            "columns": meta["columns"],
            "header_row": meta["header_row"],
            "suggested_mapping": suggested,
            "template_fields": template_field_catalog(),
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e)) from e


def get_val_by_aliases(row, columns, aliases, exclusions=None, allow_substring=True):
    import pandas as pd
    if exclusions is None:
        exclusions = []
        
    def clean_str(s):
        return re.sub(r'[\s_.-]', '', str(s).lower())
        
    clean_aliases = [clean_str(a) for a in aliases]
    clean_exclusions = [clean_str(e) for e in exclusions]
    
    # Filter columns
    filtered_cols = []
    for col in columns:
        col_clean = clean_str(col)
        if any(exc in col_clean for exc in clean_exclusions):
            continue
        filtered_cols.append(col)
        
    # Phase 1: Exact clean match (non-empty first)
    for col in filtered_cols:
        col_clean = clean_str(col)
        for ca in clean_aliases:
            if col_clean == ca:
                val = row[col]
                if pd.notna(val) and val is not None and str(val).strip():
                    return str(val).strip()
                    
    # Phase 2: Substring match (non-empty first)
    if allow_substring:
        for col in filtered_cols:
            col_clean = clean_str(col)
            for ca in clean_aliases:
                if len(ca) <= 2 and ca in ['c', 'si', 'mn', 'p', 's', 'cu', 'ni', 'mg']:
                    continue
                if ca in col_clean:
                    val = row[col]
                    if pd.notna(val) and val is not None and str(val).strip():
                        return str(val).strip()
                        
    # Phase 3: Exact clean match (allow empty)
    for col in filtered_cols:
        col_clean = clean_str(col)
        for ca in clean_aliases:
            if col_clean == ca:
                return ""
                
    # Phase 4: Substring match (allow empty)
    if allow_substring:
        for col in filtered_cols:
            col_clean = clean_str(col)
            for ca in clean_aliases:
                if len(ca) <= 2 and ca in ['c', 'si', 'mn', 'p', 's', 'cu', 'ni', 'mg']:
                    continue
                if ca in col_clean:
                    return ""
                    
    return ""


def search_metallurgy_data(heat_no, casting_name, metallurgy_path):
    try:
        if not metallurgy_path or not os.path.isfile(metallurgy_path):
            return None

        row_data = excel_service.search_heat_no(metallurgy_path, heat_no, casting_name)
        if not row_data or not row_data.get('found'):
            return None

        return {
            'customer': row_data.get('customer', ''),
            'material_grade': row_data.get('material_grade', ''),
            'casting_name': row_data.get('casting_name', ''),
            'casting_sl_no': row_data.get('casting_sl_no', ''),
            'drawing_no': row_data.get('drawing_no', ''),
            'invoice': row_data.get('invoice_no_date', ''),
            'qty': row_data.get('qty', ''),
            'chemical_actual': row_data.get('chemical_actual') or {},
            'mechanical_actual': row_data.get('mechanical_actual') or {},
        }
    except Exception as e:
        print(f"Search error: {e}")
        import traceback
        traceback.print_exc()
        return None
async def _run_search_preview(body: HeatSearchRequest) -> dict:
    import pandas as pd

    fn = _required_upload(body.metallurgy_filename, "metallurgy", "Metallurgy")
    metallurgy_path = os.path.join(UPLOADS_DIR, fn)

    mech_fn = _required_upload(body.mechanical_requirements_filename, "mechanical", "Mechanical properties requirement")
    mech_path = os.path.join(UPLOADS_DIR, mech_fn)

    if not os.path.isfile(metallurgy_path):
        raise HTTPException(status_code=404, detail="Metallurgy file not found.")

    if not os.path.isfile(mech_path):
        raise HTTPException(status_code=404, detail="Mechanical requirements file not found.")

    # 1. Search metallurgy file using search_metallurgy_data
    want_heat = str(body.heat_no).strip()
    want_casting = str(body.casting_name).strip()

    res_met = search_metallurgy_data(want_heat, want_casting, metallurgy_path)
    if not res_met:
        raise HTTPException(status_code=404, detail=f"No matching row found for Heat '{want_heat}' and Casting '{want_casting}' in metallurgy sheet.")

    customer = res_met['customer']
    material_grade = res_met['material_grade']
    casting_sl_no = res_met['casting_sl_no']
    drawing_no = res_met['drawing_no']
    invoice = res_met['invoice']
    qty = res_met.get('qty')
    chem_actual = res_met['chemical_actual']
    mech_actual = res_met['mechanical_actual']

    # Load df_metal and find match_row for printing and chemical specification parsing
    df_metal = None
    match_row = None
    casting_col_name = ' Casting Name'
    try:
        for f in os.listdir(UPLOADS_DIR):
            if 'metallurgy' in f.lower() or 'furnace' in f.lower() or 'metallurgical' in f.lower():
                df_metal = pd.read_excel(os.path.join(UPLOADS_DIR, f), header=4)
                break
        if df_metal is not None:
            heat_col = 'Heat No  '
            casting_col = ' Casting Name'
            df_metal[heat_col] = df_metal[heat_col].astype(str).str.strip()
            df_metal[casting_col] = df_metal[casting_col].astype(str).str.strip()
            
            match = df_metal[
                df_metal[heat_col].str.contains(want_heat, case=False, na=False) &
                df_metal[casting_col].str.contains(want_casting, case=False, na=False)
            ]
            if match.empty:
                match = df_metal[df_metal[heat_col].str.contains(want_heat, case=False, na=False)]
            if not match.empty:
                match_row = match.iloc[0]
    except Exception as e:
        print("Error loading df_metal/match_row:", e)

    # 2. Search mechanical requirements file
    try:
        df_mech = pd.read_excel(mech_path)
        print("Mechanical requirements file column names:", df_mech.columns.tolist())
    except Exception as e:
        print("Error reading mechanical requirements columns:", e)

    # Print all column names of both files to console
    print("\n" + "="*80)
    print("METALLURGY SHEET COLUMNS:")
    print(list(df_metal.columns))
    print("-"*80)
    print("MECHANICAL REQUIREMENTS SHEET COLUMNS:")
    try:
        print(list(df_mech.columns))
    except Exception:
        print([])
    print("="*80 + "\n")

    # User-requested printing of all metallurgy columns
    print("ALL METALLURGY COLUMNS:")
    for col in df_metal.columns:
        print(f"  {col}")

    CHEMICAL_ORDER = ["C", "Si", "Mn", "P", "S", "Cu", "Ni", "Mg"]
    from services.excel_service import get_chemical_specified
    chemical_spec_data = get_chemical_specified(material_grade)

    casting_name = want_casting
    spec_data = search_mechanical_specified(casting_name)
    print("Specified data found:", spec_data)

    return {
        "success": True,
        "basic_info": {
            "customer": customer,
            "material_grade": material_grade,
            "drawing_no": drawing_no,
            "casting_sl_no": casting_sl_no,
            "invoice": invoice,
            "qty": qty,
            "invoice_no_date": invoice,
            "heat_no": want_heat,
            "casting_name": want_casting,
            "doc_ref": "",
            "issue_no_dt": "",
            "rev_no_dt": "",
        },
        "chemical_actual": chem_actual,
        "chemical_specified": chemical_spec_data or {},
        "mechanical_actual": mech_actual,
        "mechanical_specified": spec_data
    }


@router.post("/search_heat_no")
@router.post("/search")
async def search_preview(body: HeatSearchRequest):
    return await _run_search_preview(body)


@router.post("/preview_mapped")
async def preview_mapped(body: PreviewMappedRequest):
    return await _run_search_preview(body)


@router.post("/download_report")
@router.post("/generate_report")
async def download_report(body: dict):
    import traceback
    print("Received download report request.")
    try:
        tf = _required_upload(body.get("template_filename"), "template", "Template")
        template_path = os.path.join(UPLOADS_DIR, tf)
        if not os.path.isfile(template_path):
            raise HTTPException(status_code=404, detail=f"Template file '{tf}' not found.")

        filename = f"MTR_{uuid.uuid4().hex[:10]}.xlsx"
        output_path = os.path.join(OUTPUTS_DIR, filename)
        os.makedirs(OUTPUTS_DIR, exist_ok=True)

        wb = load_workbook(template_path)
        ws = wb.active

        # Helper to set cell value with merged cell error handling
        def _set_cell_value(sheet, address, value):
            if not address:
                return
            val_str = "" if value is None else str(value).strip()
            try:
                sheet[address] = val_str
            except Exception as e:
                print(f"MergedCell/write error on cell {address}: {e}. Retrying with unmerge/write/remerge.")
                merged_range = None
                from openpyxl.utils import coordinate_to_tuple
                try:
                    row_idx, col_idx = coordinate_to_tuple(address)
                    for r in list(sheet.merged_cells.ranges):
                        if r.min_col <= col_idx <= r.max_col and r.min_row <= row_idx <= r.max_row:
                            merged_range = r
                            break
                    if merged_range:
                        range_str = str(merged_range)
                        min_col = merged_range.min_col
                        min_row = merged_range.min_row
                        sheet.unmerge_cells(range_str)
                        sheet.cell(row=min_row, column=min_col, value=val_str)
                        sheet.merge_cells(range_str)
                    else:
                        raise e
                except Exception as ex:
                    print(f"Failed to handle merged cell write for {address}: {ex}")
                    raise ex

        # Extract values from request body following FIX 3 mapping
        bi = body.get("basic_info") or {}
        
        customer = ""
        casting_name = ""
        heat_no = ""
        casting_sl_no = ""
        doc_ref = ""
        issue_no_dt = ""
        rev_no_dt = ""
        
        if hasattr(bi, "customer"):
            customer = bi.customer
            casting_name = bi.casting_name
            heat_no = bi.heat_no
            casting_sl_no = bi.casting_sl_no
            doc_ref = bi.doc_ref
            issue_no_dt = bi.issue_no_dt
            rev_no_dt = bi.rev_no_dt
        elif isinstance(bi, dict):
            customer = bi.get("customer", "")
            casting_name = bi.get("casting_name", "")
            heat_no = bi.get("heat_no", "")
            casting_sl_no = bi.get("casting_sl_no", "")
            doc_ref = bi.get("doc_ref", "")
            issue_no_dt = bi.get("issue_no_dt", "")
            rev_no_dt = bi.get("rev_no_dt", "")

        # Fallback to top-level attributes if not set
        if not customer: customer = body.get("customer", "")
        if not casting_name: casting_name = body.get("casting_name", "")
        if not heat_no: heat_no = body.get("heat_no", "")
        if not casting_sl_no: casting_sl_no = body.get("casting_sl_no", "")
        if not doc_ref: doc_ref = body.get("doc_ref", "")
        if not issue_no_dt: issue_no_dt = body.get("issue_no_dt", "")
        if not rev_no_dt: rev_no_dt = body.get("rev_no_dt", "")

        def _get_bi_val(b_info, main_body, key, default=""):
            if hasattr(b_info, key): return getattr(b_info, key) or default
            if isinstance(b_info, dict): return b_info.get(key, default) or default
            return main_body.get(key, default) or default

        material_grade = _get_bi_val(bi, body, "material_grade")
        drawing_no = _get_bi_val(bi, body, "drawing_no")
        invoice_no_date = _get_bi_val(bi, body, "invoice_no_date")
        date_val = _get_bi_val(bi, body, "date")

        # Chemical actual row 19
        chem_act = body.get("chemical_actual") or {}
        chem_actual = chem_act  # alias used below
        c_actual = chem_act.get("C", "")
        si_actual = chem_act.get("Si", "")
        mn_actual = chem_act.get("Mn", "")
        p_actual = chem_act.get("P", "")
        s_actual = chem_act.get("S", "")
        cu_actual = chem_act.get("Cu", "")
        ni_actual = chem_act.get("Ni", "")
        mg_actual = chem_act.get("Mg", "")

        # Chemical specified row 17 - looked up by material grade
        from services.excel_service import get_chemical_specified
        chem_spec = get_chemical_specified(material_grade) or {}

        # Mechanical specified row 27
        mech_spec = body.get("mechanical_specified") or {}
        tensile_spec = mech_spec.get("tensile", "")
        proof_spec = mech_spec.get("proof_stress", "")
        elongation_spec = mech_spec.get("elongation", "")
        hardness_spec = mech_spec.get("hardness", "") or mech_spec.get("hardness_bhn", "")
        impact_ind_spec = mech_spec.get("impact_individual", "")
        impact_mean_spec = mech_spec.get("impact_mean", "")

        # Mechanical actual row 29 & 30
        mech_act = body.get("mechanical_actual") or {}
        tensile_actual = mech_act.get("tensile", "")
        proof_actual = mech_act.get("proof_stress", "")
        elongation_actual = mech_act.get("elongation", "")
        hardness_actual = mech_act.get("hardness", "") or mech_act.get("hardness_bhn", "")
        impact_ind_actual = mech_act.get("impact_individual", "")
        impact_mean_actual = mech_act.get("impact_mean", "")

        # Step 1: scan uploads for template file (any file with 'template' in name)
        template_path_found = None
        for f in os.listdir(UPLOADS_DIR):
            if 'template' in f.lower():
                template_path_found = os.path.join(UPLOADS_DIR, f)
                break
        if template_path_found:
            template_path = template_path_found

        # Step 2: load template
        from openpyxl import load_workbook as _lw
        wb2 = _lw(template_path)
        sheet = wb2.active
        _replace_cover_logo(sheet, os.path.join(BASE_DIR, "company_logo.png"))
        _clear_cover_right_block(sheet)

        def safe_write(sheet, cell, value):
            if value is None:
                return
            val_str = str(value).strip().lower()
            if val_str == "" or val_str == "none" or val_str == "nan":
                return
            
            try:
                sheet[cell] = value
            except:
                for mr in list(sheet.merged_cells.ranges):
                    if cell in mr:
                        sheet.unmerge_cells(str(mr))
                        sheet[cell] = value
                        sheet.merge_cells(str(mr))
                        break

        # Step 4: write all cells
        safe_write(sheet, 'D6', customer)
        safe_write(sheet, 'K6', material_grade)
        safe_write(sheet, 'D8', casting_name)
        safe_write(sheet, 'K8', drawing_no)
        safe_write(sheet, 'D10', heat_no)
        safe_write(sheet, 'K10', casting_sl_no)
        
        invoice_val = _get_bi_val(bi, body, "invoice")
        qty_val = _get_bi_val(bi, body, "qty")
        inv_qty_str = f"{invoice_val} / {qty_val}" if invoice_val and qty_val else f"{invoice_val or ''}{qty_val or ''}"
        safe_write(sheet, 'K12', inv_qty_str)

        safe_write(sheet,'C17', chem_spec.get('C'))
        safe_write(sheet,'D17', chem_spec.get('Si'))
        safe_write(sheet,'E17', chem_spec.get('Mn'))
        safe_write(sheet,'F17', chem_spec.get('P'))
        safe_write(sheet,'G17', chem_spec.get('S'))
        safe_write(sheet,'H17', chem_spec.get('Cu'))
        safe_write(sheet,'I17', chem_spec.get('Ni'))
        safe_write(sheet,'J17', chem_spec.get('Mg'))
        
        safe_write(sheet,'C19', chem_actual.get('C'))
        safe_write(sheet,'D19', chem_actual.get('Si'))
        safe_write(sheet,'E19', chem_actual.get('Mn'))
        safe_write(sheet,'F19', chem_actual.get('P'))
        safe_write(sheet,'G19', chem_actual.get('S'))
        safe_write(sheet,'H19', chem_actual.get('Cu'))
        safe_write(sheet,'I19', chem_actual.get('Ni'))
        safe_write(sheet,'J19', chem_actual.get('Mg'))

        safe_write(sheet,'D27', tensile_spec)
        safe_write(sheet,'E27', proof_spec)
        safe_write(sheet,'F27', elongation_spec)
        safe_write(sheet,'H27', hardness_spec)
        safe_write(sheet,'J27', impact_ind_spec)
        safe_write(sheet,'L27', impact_mean_spec)

        safe_write(sheet,'D29', tensile_actual)
        safe_write(sheet,'E29', proof_actual)
        safe_write(sheet,'F29', elongation_actual)
        safe_write(sheet,'H29', hardness_actual)

        impact_ind_1 = mech_act.get("impact_individual_1", "")
        impact_ind_2 = mech_act.get("impact_individual_2", "")
        impact_ind_3 = mech_act.get("impact_individual_3", "")

        def format_impact_values(v1, v2, v3):
            vals = [str(x).strip() for x in [v1, v2, v3] if str(x).strip() and str(x).strip() != 'None']
            if len(vals) > 1:
                return ", ".join(vals)
            if len(vals) == 1:
                val = vals[0]
                if val.isdigit() and len(val) >= 4 and len(val) % 2 == 0:
                    return ", ".join([val[i:i+2] for i in range(0, len(val), 2)])
                return val
            return ""

        impact_combined = format_impact_values(impact_ind_1, impact_ind_2, impact_ind_3)
        safe_write(sheet, 'J29', impact_combined)
        safe_write(sheet, 'J30', "")
        safe_write(sheet, 'J31', "")
        safe_write(sheet,'L30', impact_mean_actual)

        # Fix 4: Preserve plus minus symbol in row 22
        for cell in sheet[22]:
            if cell.value and isinstance(cell.value, str):
                if '-20' in cell.value or '20' in cell.value:
                    cell.value = "at -20±1° C"

        # Step 5: save and return
        os.makedirs(OUTPUTS_DIR, exist_ok=True)
        filename = f"Test_Report_{heat_no or 'test'}.xlsx"
        output = os.path.join(OUTPUTS_DIR, filename)
        wb2.save(output)
        print(f"Report saved successfully to {output}")

        return FileResponse(
            path=output,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print("ERROR IN DOWNLOAD ENDPOINT:")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e)) from e

